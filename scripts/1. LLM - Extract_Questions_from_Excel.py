#%%
import openpyxl
import json
from pathlib import Path

# Define paths
script_dir = Path(__file__).parent
project_root = script_dir.parent
excel_path = project_root / "data" / "DRM System Diagnostic Assessment - Template.xlsx"
output_dir = project_root / "LLM"
output_file = output_dir / "questions_mapping.json"

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_path, data_only=True)
sheet = workbook.active

# Dictionary to store question mappings - nested by thematic area
questions_dict = {}

# Track current thematic area (Column C has merged cells)
current_thematic_area = None
current_pillar = None
question_counter = 1  # Global question counter

# Iterate through rows starting from row 15 (first data row after headers)
for row_num in range(15, 100):
    pillar = sheet.cell(row_num, 2).value  # Column B - Pillar
    thematic = sheet.cell(row_num, 3).value  # Column C - Thematic Area (merged)
    question_text = sheet.cell(row_num, 4).value  # Column D - Question
    print(pillar, thematic)

    # Update current pillar if found
    if pillar:
        current_pillar = pillar.strip()
        # If pillar is updated but thematic is not, use pillar as thematic area
        if not thematic:
            current_thematic_area = current_pillar
    
    # Update current thematic area if found (handles merged cells)
    if thematic:
        current_thematic_area = thematic.strip()
    
    # Process question if we have all required data
    if question_text and current_thematic_area:
        question_text = question_text.strip()
        question_id = f"Q{question_counter}"
        
        # Initialize thematic area in dict if not exists
        if current_thematic_area not in questions_dict:
            questions_dict[current_thematic_area] = {}
        
        # Add question under thematic area
        questions_dict[current_thematic_area][question_id] = question_text
        
        print(f"Extracted {question_id}: {current_thematic_area} - {question_text[:50]}...")
        question_counter += 1
    
    # Stop if we've gone past the data
    if question_text is None and pillar is None and thematic is None:
        # Check if we have at least 3 consecutive empty rows
        if (sheet.cell(row_num + 1, 4).value is None and 
            sheet.cell(row_num + 2, 4).value is None):
            break

# Save to JSON file
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(questions_dict, f, indent=2, ensure_ascii=False)

print(f"\nSuccessfully extracted {question_counter - 1} questions across {len(questions_dict)} thematic areas")
print(f"JSON file saved to: {output_file}")

# Display sample output
print(f"\nSample output:")
for thematic_area, questions in list(questions_dict.items())[:2]:
    print(f'\n{thematic_area}:')
    for q_id, q_text in list(questions.items())[:2]:
        print(f'  {q_id}: "{q_text[:60]}..."')